"""
Thames Path Status — weekly auto-updater
"""

import os
import json
import time
from datetime import date
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

TODAY     = date.today().isoformat()
XLSX_PATH = "PathStatus.xlsx"
CSV_PATH  = "PathStatus.csv"
LOG_PATH  = "update_log.txt"

# ── Prompts ───────────────────────────────────────────────────────────────────

SYSTEM_PROMPT = f"""You research Thames Path closures and return ONLY a JSON array.
No markdown, no preamble, no explanation. Just the raw JSON array starting with [.
Today: {TODAY}.

Each item must have these exact fields:
type (closure/construction/incident/info),
type2 (diversion/intermittent/open/closure/null),
title (max 80 chars),
description (max 400 chars),
mile (float or null),
status (short string),
date (YYYY-MM-DD or null),
last_verified ("{TODAY}"),
source (URL or null),
lat (float or null),
lon (float or null)"""

USER_PROMPT = """Search for current Thames Path closures and diversions.
Check nationaltrail.co.uk and walkthethames.co.uk/thames-path-status

Verify these known issues (confirm still active or mark reopened):
- Osney Bridge diversion Oxford mile 53.8
- Marsh Lock closure Henley mile 105.1
- Sandford Footbridge closure mile 57
- Abingdon Weir intermittent closure mile 62
- Temple Footbridge closure near Marlow mile 108.5
- Runnymede Bridge 142 closure Egham mile 136.5
- Bell Weir Lock diversion Runnymede mile 136.2
- Streatley-Goring towpath diversion mile 81.5
- Penton Hook Island closure mile 135.6
- Ten Foot Bridge closure Faringdon mile 34
- Brentford Grand Union Canal disruption mile 154

Return ONLY a valid JSON array, nothing else."""

# ── Claude API ────────────────────────────────────────────────────────────────

def fetch_updates_from_claude():
    client = anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])
    print("Calling Claude API...")

    messages = [{"role": "user", "content": USER_PROMPT}]
    tools = [{"type": "web_search_20250305", "name": "web_search"}]

    # Agentic loop — keep going until we get a final text response
    full_text = ""
    for attempt in range(10):
        response = client.messages.create(
            model="claude-haiku-4-5",
            max_tokens=2000,
            system=SYSTEM_PROMPT,
            tools=tools,
            messages=messages,
        )

        print(f"  Turn {attempt+1}: stop_reason={response.stop_reason}, blocks={[b.type for b in response.content]}")

        if response.stop_reason == "end_turn":
            text_blocks = [b.text for b in response.content if hasattr(b, "text") and b.text.strip()]
            if text_blocks:
                full_text = text_blocks[-1].strip()
                print(f"  Got text response ({len(full_text)} chars)")
                break
            else:
                raise ValueError("end_turn but no text content in response")

        elif response.stop_reason == "tool_use":
            messages.append({"role": "assistant", "content": response.content})
            tool_results = []
            for block in response.content:
                if block.type == "tool_use":
                    tool_results.append({
                        "type": "tool_result",
                        "tool_use_id": block.id,
                        "content": "Search completed successfully.",
                    })
            messages.append({"role": "user", "content": tool_results})
            time.sleep(1)

        else:
            raise ValueError(f"Unexpected stop_reason: {response.stop_reason}")
    else:
        raise ValueError("Exceeded maximum turns without getting a final response")

    # Strip any accidental markdown fences
    if full_text.startswith("```"):
        full_text = full_text.split("\n", 1)[1].rsplit("```", 1)[0]

    # Find the JSON array in the response
    start = full_text.find("[")
    end   = full_text.rfind("]") + 1
    if start == -1 or end == 0:
        raise ValueError(f"No JSON array found in response: {full_text[:200]}")

    return json.loads(full_text[start:end])


# ── Write Excel ───────────────────────────────────────────────────────────────

HEADERS    = ["Type","Type2","Title","Description","Mile","Status",
              "Date","Last Verified","Source","Lat","Lon"]
COL_WIDTHS = [14,14,45,80,6,28,12,14,60,12,12]

def write_xlsx(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PathStatus"

    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    dfont = Font(name="Arial", size=10)
    wrap  = Alignment(wrap_text=True, vertical="top")
    ctr   = Alignment(horizontal="center", vertical="center")

    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, c, h)
        cell.font = hfont; cell.fill = hfill; cell.alignment = ctr
    ws.row_dimensions[1].height = 20

    for r, row in enumerate(rows, 2):
        vals = [row.get("type"), row.get("type2"), row.get("title"),
                row.get("description"), row.get("mile"), row.get("status"),
                row.get("date"), row.get("last_verified", TODAY),
                row.get("source"), row.get("lat"), row.get("lon")]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(r, c, val)
            cell.font = dfont; cell.alignment = wrap
        ws.row_dimensions[r].height = 60

    for c, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[ws.cell(1,c).column_letter].width = w
    ws.freeze_panes = "A2"
    wb.save(XLSX_PATH)
    print(f"Saved {len(rows)} rows to {XLSX_PATH}")


# ── Write CSV ─────────────────────────────────────────────────────────────────

def write_csv(rows):
    import csv
    with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "Type","Type2","Title","Description","Mile","Status",
            "Date","Last Verified","Source","Lat","Lon"])
        writer.writeheader()
        for row in rows:
            writer.writerow({
                "Type": row.get("type",""),
                "Type2": row.get("type2",""),
                "Title": row.get("title",""),
                "Description": row.get("description",""),
                "Mile": row.get("mile",""),
                "Status": row.get("status",""),
                "Date": row.get("date",""),
                "Last Verified": row.get("last_verified", TODAY),
                "Source": row.get("source",""),
                "Lat": row.get("lat",""),
                "Lon": row.get("lon",""),
            })
    print(f"Saved CSV to {CSV_PATH}")


# ── Log ───────────────────────────────────────────────────────────────────────

def write_log(rows, error=None):
    with open(LOG_PATH, "w") as f:
        f.write(f"Thames Path Update Log\nDate: {TODAY}\n")
        if error:
            f.write(f"ERROR: {error}\n")
        else:
            f.write(f"Status: OK\nRows: {len(rows)}\n\n")
            for row in rows:
                f.write(f"  [{row.get('status','?')}] {row.get('title','?')}\n")


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    try:
        rows = fetch_updates_from_claude()
        print(f"Received {len(rows)} rows")
        write_xlsx(rows)
        write_csv(rows)
        write_log(rows)
        print("Done.")
    except Exception as e:
        print(f"ERROR: {e}")
        write_log([], error=str(e))
        raise
